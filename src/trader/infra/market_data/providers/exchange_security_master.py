"""Atomic free security-master snapshot from the official SSE and SZSE lists."""

from __future__ import annotations

import hashlib
import threading
import time
import warnings
from collections.abc import Callable, Mapping, Sequence
from dataclasses import dataclass
from datetime import date, datetime
from datetime import time as datetime_time
from io import BytesIO
from typing import TypedDict, cast
from zoneinfo import ZoneInfo

import requests
from typing_extensions import Unpack

from trader.application.cache import canonical_json_bytes
from trader.infra.market_data.service.observations import SourceObservation

_SHANGHAI = ZoneInfo("Asia/Shanghai")
_SSE_URL = "https://query.sse.com.cn/sseQuery/commonQuery.do"
_SZSE_URL = "https://www.szse.cn/api/report/ShowReport"
_SOURCE = "exchange_security_master"
_CONTRACT_VERSION = "exchange-security-master-v1"
_SUPPORTED_BOARDS = frozenset({"main", "chinext", "star"})
_SUPPORTED_EXCHANGES = frozenset({"SSE", "SZSE"})


@dataclass(frozen=True)
class ExchangeSecurityMasterListing:
    code: str
    name: str
    listing_date: date
    board: str
    exchange: str


@dataclass(frozen=True)
class ExchangeSecurityMasterHealthStatus:
    enabled: bool
    planned_count: int
    success_count: int
    error_count: int
    timeout_count: int
    consecutive_failures: int
    last_latency_ms: float | None
    p50_latency_ms: float | None
    p95_latency_ms: float | None
    last_error: str | None
    snapshot_rows: int
    listing_date_rows: int
    last_source_time: datetime | None
    timeout_seconds: float


ListingFetcher = Callable[[float], Sequence[ExchangeSecurityMasterListing]]


class _RequiredClientOptions(TypedDict):
    timeout_seconds: float
    wall_clock: Callable[[], datetime]


class _OptionalClientOptions(TypedDict, total=False):
    minimum_rows: int
    sse_fetcher: ListingFetcher
    szse_fetcher: ListingFetcher
    monotonic: Callable[[], float]


class _ClientOptions(_RequiredClientOptions, _OptionalClientOptions):
    pass


class _ExchangeSourceError(RuntimeError):
    def __init__(self, source: str, cause: Exception) -> None:
        super().__init__(f"{source} security master fetch failed")
        self.source = source
        self.cause = cause


class ExchangeSecurityMasterClient:
    """Fetch and validate one all-or-nothing supported A-share identity snapshot."""

    def __init__(self, **options: Unpack[_ClientOptions]) -> None:
        timeout_seconds = options["timeout_seconds"]
        minimum_rows = options.get("minimum_rows", 4_000)
        if timeout_seconds <= 0.0:
            raise ValueError("exchange security master timeout must be positive")
        if minimum_rows <= 0:
            raise ValueError("exchange security master minimum rows must be positive")
        self._timeout_seconds = float(timeout_seconds)
        self._minimum_rows = minimum_rows
        self._sse_fetcher = options.get("sse_fetcher", _fetch_sse_listings)
        self._szse_fetcher = options.get("szse_fetcher", _fetch_szse_listings)
        self._monotonic = options.get("monotonic", time.monotonic)
        self._wall_clock = options["wall_clock"]
        self._lock = threading.Lock()
        self._planned_count = 0
        self._success_count = 0
        self._error_count = 0
        self._timeout_count = 0
        self._consecutive_failures = 0
        self._latencies_ms: list[float] = []
        self._last_error: str | None = None
        self._snapshot_rows = 0
        self._listing_date_rows = 0
        self._last_source_time: datetime | None = None

    def fetch(self, observed_at: datetime) -> tuple[SourceObservation, ...]:
        if observed_at.tzinfo is None or observed_at.utcoffset() is None:
            raise ValueError("exchange security master observed_at must be timezone-aware")
        started = self._monotonic()
        with self._lock:
            self._planned_count += 1
        try:
            try:
                sse_listings = tuple(self._sse_fetcher(self._timeout_seconds))
            except Exception as exc:
                raise _ExchangeSourceError("sse", exc) from exc
            try:
                szse_listings = tuple(self._szse_fetcher(self._timeout_seconds))
            except Exception as exc:
                raise _ExchangeSourceError("szse", exc) from exc
            listings = sse_listings + szse_listings
            normalized = _validate_snapshot(listings, self._minimum_rows, observed_at)
            received_at = self._wall_clock()
            if received_at.tzinfo is None or received_at.utcoffset() is None:
                raise ValueError("exchange security master wall clock must be timezone-aware")
            version = _snapshot_version(normalized)
            observations = tuple(
                _to_observation(listing, observed_at=observed_at, received_at=received_at, version=version)
                for listing in normalized
            )
        except Exception as exc:
            self._record_failure(exc, started)
            raise
        self._record_success(len(observations), received_at, started)
        return observations

    def health(self) -> ExchangeSecurityMasterHealthStatus:
        with self._lock:
            latencies = tuple(self._latencies_ms)
            return ExchangeSecurityMasterHealthStatus(
                enabled=True,
                planned_count=self._planned_count,
                success_count=self._success_count,
                error_count=self._error_count,
                timeout_count=self._timeout_count,
                consecutive_failures=self._consecutive_failures,
                last_latency_ms=round(latencies[-1], 2) if latencies else None,
                p50_latency_ms=_percentile(latencies, 0.50),
                p95_latency_ms=_percentile(latencies, 0.95),
                last_error=self._last_error,
                snapshot_rows=self._snapshot_rows,
                listing_date_rows=self._listing_date_rows,
                last_source_time=self._last_source_time,
                timeout_seconds=self._timeout_seconds,
            )

    def _record_success(self, rows: int, source_time: datetime, started: float) -> None:
        with self._lock:
            self._success_count += 1
            self._consecutive_failures = 0
            self._last_error = None
            self._snapshot_rows = rows
            self._listing_date_rows = rows
            self._last_source_time = source_time
            self._record_latency_locked(started)

    def _record_failure(self, exc: Exception, started: float) -> None:
        error_code = _error_code(exc)
        with self._lock:
            self._error_count += 1
            self._consecutive_failures += 1
            self._timeout_count += int(error_code == "timeout" or error_code.endswith("_timeout"))
            self._last_error = error_code
            self._record_latency_locked(started)

    def _record_latency_locked(self, started: float) -> None:
        self._latencies_ms.append(max(0.0, (self._monotonic() - started) * 1_000.0))
        del self._latencies_ms[:-128]


def _validate_snapshot(
    listings: Sequence[ExchangeSecurityMasterListing],
    minimum_rows: int,
    observed_at: datetime,
) -> tuple[ExchangeSecurityMasterListing, ...]:
    by_code: dict[str, ExchangeSecurityMasterListing] = {}
    local_date = observed_at.astimezone(_SHANGHAI).date()
    invalid = False
    for listing in listings:
        if (
            len(listing.code) != 6
            or not listing.code.isdigit()
            or not listing.name.strip()
            or listing.board not in _SUPPORTED_BOARDS
            or listing.exchange not in _SUPPORTED_EXCHANGES
            or listing.listing_date > local_date
        ):
            invalid = True
            continue
        previous = by_code.get(listing.code)
        if previous is not None:
            invalid = True
        by_code[listing.code] = listing
    exchanges = {listing.exchange for listing in by_code.values()}
    if invalid or len(by_code) < minimum_rows or exchanges != _SUPPORTED_EXCHANGES:
        raise ValueError("exchange security master snapshot is incomplete or conflicting")
    return tuple(by_code[code] for code in sorted(by_code))


def _snapshot_version(listings: Sequence[ExchangeSecurityMasterListing]) -> str:
    payload = [
        {
            "code": item.code,
            "name": item.name,
            "listing_date": item.listing_date.isoformat(),
            "board": item.board,
            "exchange": item.exchange,
        }
        for item in listings
    ]
    return f"{_CONTRACT_VERSION}:{hashlib.sha256(canonical_json_bytes(payload)).hexdigest()}"


def _to_observation(
    listing: ExchangeSecurityMasterListing,
    *,
    observed_at: datetime,
    received_at: datetime,
    version: str,
) -> SourceObservation:
    fields = {
        "board": listing.board,
        "board_reliability": "verified",
        "exchange": listing.exchange,
        "listing_date": listing.listing_date.isoformat(),
    }
    return SourceObservation(
        source=_SOURCE,
        subject_key=listing.code,
        observed_at=observed_at,
        source_time=received_at,
        received_at=received_at,
        effective_at=datetime.combine(listing.listing_date, datetime_time.min, _SHANGHAI),
        data_version=version,
        fields=fields,
        missing_reasons={},
        payload_hash=hashlib.sha256(canonical_json_bytes(fields)).hexdigest(),
        status="success",
        error_code=None,
    )


def _fetch_sse_listings(timeout_seconds: float) -> tuple[ExchangeSecurityMasterListing, ...]:
    headers = {
        "Host": "query.sse.com.cn",
        "Pragma": "no-cache",
        "Referer": "https://www.sse.com.cn/assortment/stock/list/share/",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36"
        ),
    }
    listings: list[ExchangeSecurityMasterListing] = []
    for stock_type, board in (("1", "main"), ("8", "star")):
        response = _official_get(
            _SSE_URL,
            params={
                "STOCK_TYPE": stock_type,
                "REG_PROVINCE": "",
                "CSRC_CODE": "",
                "STOCK_CODE": "",
                "sqlId": "COMMON_SSE_CP_GPJCTPZ_GPLB_GP_L",
                "COMPANY_STATUS": "2,4,5,7,8",
                "type": "inParams",
                "isPagination": "true",
                "pageHelp.cacheSize": "1",
                "pageHelp.beginPage": "1",
                "pageHelp.pageSize": "10000",
                "pageHelp.pageNo": "1",
                "pageHelp.endPage": "1",
            },
            headers=headers,
            timeout=timeout_seconds,
        )
        try:
            payload = response.json()
        finally:
            response.close()
        if not isinstance(payload, Mapping) or not isinstance(payload.get("result"), list):
            raise ValueError("SSE security master response shape is invalid")
        for raw in cast(list[object], payload["result"]):
            if not isinstance(raw, Mapping):
                raise ValueError("SSE security master row shape is invalid")
            listings.append(
                ExchangeSecurityMasterListing(
                    code=_code(raw.get("A_STOCK_CODE")),
                    name=str(raw.get("SEC_NAME_CN") or "").strip(),
                    listing_date=_date_value(raw.get("LIST_DATE")),
                    board=board,
                    exchange="SSE",
                )
            )
    return tuple(listings)


def _fetch_szse_listings(timeout_seconds: float) -> tuple[ExchangeSecurityMasterListing, ...]:
    response = _official_get(
        _SZSE_URL,
        params={
            "SHOWTYPE": "xlsx",
            "CATALOGID": "1110",
            "TABKEY": "tab1",
            "random": f"{time.time():.6f}",
        },
        headers={
            "Referer": "https://www.szse.cn/market/product/stock/list/index.html",
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36"
            ),
        },
        timeout=timeout_seconds,
    )
    try:
        content = response.content
    finally:
        response.close()
    from openpyxl import load_workbook

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = load_workbook(BytesIO(content), read_only=False, data_only=True)
    try:
        rows = workbook.active.iter_rows(values_only=True)
        try:
            headers = tuple(str(value).strip() if value is not None else "" for value in next(rows))
        except StopIteration as exc:
            raise ValueError("SZSE security master workbook is empty") from exc
        required = ("板块", "A股代码", "A股简称", "A股上市日期")
        try:
            indices = {name: headers.index(name) for name in required}
        except ValueError as exc:
            raise ValueError("SZSE security master workbook columns are invalid") from exc
        listings: list[ExchangeSecurityMasterListing] = []
        for row in rows:
            code = _code(row[indices["A股代码"]])
            if not code:
                continue
            board = _szse_board(row[indices["板块"]])
            if board is None:
                continue
            listings.append(
                ExchangeSecurityMasterListing(
                    code=code,
                    name=str(row[indices["A股简称"]] or "").strip(),
                    listing_date=_date_value(row[indices["A股上市日期"]]),
                    board=board,
                    exchange="SZSE",
                )
            )
    finally:
        workbook.close()
    return tuple(listings)


def _official_get(
    url: str,
    *,
    params: Mapping[str, str],
    timeout: float,
    headers: Mapping[str, str] | None = None,
) -> requests.Response:
    last_error: Exception | None = None
    retry_delays = (1.0, 3.0, 5.0)
    for attempt in range(4):
        try:
            response = requests.get(url, params=params, headers=headers, timeout=timeout)
            response.raise_for_status()
            return response
        except (requests.ConnectionError, requests.Timeout) as exc:
            last_error = exc
        except requests.HTTPError as exc:
            if exc.response is None or exc.response.status_code < 500:
                raise
            last_error = exc
        if attempt < len(retry_delays):
            time.sleep(retry_delays[attempt])
    assert last_error is not None
    raise last_error


def _code(value: object) -> str:
    if isinstance(value, bool) or value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(int(value)).zfill(6)
    text = str(value).strip().partition(".")[0]
    return text.zfill(6) if text.isdigit() else ""


def _date_value(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip().replace("/", "-")
    if len(text) == 8 and text.isdigit():
        text = f"{text[:4]}-{text[4:6]}-{text[6:]}"
    return date.fromisoformat(text)


def _szse_board(value: object) -> str | None:
    text = str(value or "").strip()
    if "创业板" in text:
        return "chinext"
    if "主板" in text:
        return "main"
    return None


def _error_code(exc: Exception) -> str:
    if isinstance(exc, _ExchangeSourceError):
        return f"{exc.source}_{_error_code(exc.cause)}"
    if isinstance(exc, (requests.Timeout, TimeoutError)):
        return "timeout"
    if isinstance(exc, ValueError):
        return "invalid_snapshot"
    if isinstance(exc, (requests.ConnectionError, OSError)):
        return "connection_failed"
    if isinstance(exc, requests.HTTPError):
        return "http_error"
    return "adapter_error"


def _percentile(values: Sequence[float], quantile: float) -> float | None:
    if not values:
        return None
    ordered = sorted(values)
    index = max(0, min(len(ordered) - 1, int(len(ordered) * quantile + 0.999999) - 1))
    return round(ordered[index], 2)


__all__ = [
    "ExchangeSecurityMasterClient",
    "ExchangeSecurityMasterHealthStatus",
    "ExchangeSecurityMasterListing",
]
